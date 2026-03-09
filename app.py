from flask import Flask, render_template, request
from openpyxl import Workbook, load_workbook
from datetime import datetime
import os

app = Flask(__name__)

FILE_PATH = "survey_results.xlsx"


def save_to_excel(data):

    headers = [
        "日時",
        "店",
        "部門",
        "名前",

        "Q1 週休2日",
        "Q2 固定残業",

        "固定残業ありの場合",
        "固定残業なしの場合",

        "Q3 サービス残業時間",

        "Q4 サービス残業理由",
        "Q4 理由その他",

        "Q5 上司相談",
        "Q6 相談結果",
        "Q7 相談しない理由",

        "Q8 有給",
        "Q9 有給理由",

        "Q10 ハラスメント",
        "ハラスメント内容",
        "ハラスメント相談",
        "ハラスメント結果",

        "Q11 ハラスメント目撃",
        "Q12 ハラスメント会社対策",

        "会社要望",
        "組合要望"
    ]

    if not os.path.exists(FILE_PATH):
        wb = Workbook()
        ws = wb.active
        ws.append(headers)
        wb.save(FILE_PATH)

    wb = load_workbook(FILE_PATH)
    ws = wb.active

    ws.append(data)
    wb.save(FILE_PATH)


@app.route("/survey")
def survey():
    return render_template("survey.html")


@app.route("/submit", methods=["POST"])
def submit():

    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    store = request.form.get("store", "")
    dept = request.form.get("dept", "")
    name = request.form.get("name", "")

    q1 = request.form.get("q1", "")
    overtime = request.form.get("overtime", "")

    q3 = request.form.get("q3", "")
    q4 = request.form.get("q4", "")

    service = request.form.get("service", "")

    reason = request.form.get("reason", "")
    reason_other = request.form.get("reason_other", "")

    consult = request.form.get("consult", "")
    solve = request.form.get("solve", "")
    why = request.form.get("why", "")

    vacation = request.form.get("vacation", "")
    vacation_reason = request.form.get("vacation_reason", "")

    harassment = request.form.get("harassment", "")
    harassment_content = request.form.get("harassment_content", "")
    harassment_consult = request.form.get("harassment_consult", "")
    harassment_result = request.form.get("harassment_result", "")

    harassment_seen = request.form.get("harassment_seen", "")
    company_measure = request.form.get("company_measure", "")

    company_request = request.form.get("company_request", "")
    union_request = request.form.get("union_request", "")

    row = [
        now,
        store,
        dept,
        name,

        q1,
        overtime,

        q3,
        q4,

        service,

        reason,
        reason_other,

        consult,
        solve,
        why,

        vacation,
        vacation_reason,

        harassment,
        harassment_content,
        harassment_consult,
        harassment_result,

        harassment_seen,
        company_measure,

        company_request,
        union_request
    ]

    save_to_excel(row)

    return render_template("thanks.html")


if __name__ == "__main__":
    app.run()